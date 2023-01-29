import openpyxl
import json

class TimeTableParser:

    def __init__(self, timetable : str):
        self.wb = openpyxl.load_workbook(timetable)
        self.ws = [i for i, sheet in enumerate(self.wb.sheetnames) if "tt" in sheet.lower()][0]
        self.ws = self.wb.worksheets[self.ws]

    def __skip_to_next_value_col(self, row : int, col : int) -> tuple[int, int]:
        cell = self.ws.cell(row=row, column=col)

        while cell.value == None:
            col += 1
            cell = self.ws.cell(row=row, column=col)

        return row, col

    def get_codes(self) -> set:
        row = 2
        col = 1
        cell = self.ws.cell(row=row, column=col)
        codes = set()

        while cell.value.lower() != 'color codes':
            col += 1
            cell = self.ws.cell(row=row, column=col)

        row, col = self.__skip_to_next_value_col(row, col + 1)

        for col in range(col, 100):
            cell = self.ws.cell(row=row, column=col)
            
            if cell.value == None:
                continue

            if cell.value.split("-")[-1].strip().isnumeric():
                codes.add(cell.value.split("-")[0].strip().lower())


        return codes
        
    def get_hour(self, row : int, col : int) -> int:
        
        cell = self.ws.cell(row=row, column=col)

        while cell.value == None:
            col -= 1
            cell = self.ws.cell(row=row, column=col)

        return int(cell.value.split(":")[0].strip()), cell.value.strip().split(" ")[-1].strip()
    
    def start_time(self, timerow : int, col : int) -> str:
        minutes = self.ws.cell(row=timerow, column=col).value
        minutes = int(minutes) - 10

        hour, timestamp = self.get_hour(timerow + 1, col)

        if timestamp == "noon":
            timestamp = "p.m"

        minutes = minutes if minutes > 9 else f"0{minutes}"

        return f"{hour}:{minutes} {timestamp.strip('.')}"
    
    def end_time(self, timerow : int, row : int, col : int) -> str:
        col += 1
        cell = self.ws.cell(row=row, column=col)
        while int(cell.font.sz) == 11:
            col += 1
            cell = self.ws.cell(row=row, column=col)

        minutes = self.ws.cell(row=timerow, column=col).value
        minutes = int(minutes)

        hour, timestamp = self.get_hour(timerow + 1, col)

        if timestamp == "noon":
            timestamp = "p.m"

        if minutes >= 60:
            minutes -= 60
            if hour == 11:
                timestamp = "p.m"
                hour += 1
            elif hour == 12:
                hour = 1
            else:
                hour += 1

        minutes = minutes if minutes > 9 else f"0{minutes}"

        return f"{hour}:{minutes} {timestamp.strip('.')}"
        
    def get_day(self, row : int, days_col : int) -> str:

        cell = self.ws.cell(row=row, column=days_col)
        while cell.value == None:
            row -= 1
            cell = self.ws.cell(row=row, column=days_col)

        return cell.value.strip().split(" ")[0].strip()

    def parse(self) -> dict:
        codes = self.get_codes()
        courses = {"timetable" : dict(), "timings" : set()}

        time_row = 2
        col = 1

        for row in range(time_row, 100):
            cell = self.ws.cell(row=row, column=col)
            if cell.value.lower().strip() == "periods":
                time_row = row
                break

        days_col = 1
        venue_col = 1
        row = time_row + 1

        for col in range(days_col, 50):
            cell = self.ws.cell(row=row, column=col)
            if cell.value.lower().strip() == "days":
                days_col = col
            if cell.value.lower().strip() == "room":
                venue_col = col
                break
        
        row = time_row + 2
        
        for row in range(row, 400):
            for col in range(1, 100):
                cell = self.ws.cell(row=row, column=col)

                if cell.value == None:
                    continue

                if cell.value.split("(")[-1].split("-")[0].lower() in codes:
                    coursename = cell.value.split("(")[0].strip().lower()
                    courses["timetable"].setdefault(coursename, dict())
                    section = cell.value.split("(")[-1].split(")")[0].strip().lower()

                    courses["timetable"][coursename].setdefault(section, list())

                    start_time = self.start_time(time_row, col)
                    end_time = self.end_time(time_row, row, col)

                    info_list = [
                        self.get_day(row, days_col),
                        self.ws.cell(row=row, column=venue_col).value.strip(),
                        start_time,
                        end_time
                    ]

                    courses["timetable"][coursename][section].append(info_list)
                    
                    courses["timings"].add(start_time)
                    courses["timings"].add(end_time)
                    
        courses["timings"] = list(courses["timings"])
        courses["timings"] = sorted(courses["timings"], key=lambda v : 
            int(v.split(" ")[0].replace(":", "")) if v.split(" ")[-1] == "a.m" or v.split(":")[0] == "12"
            else int(v.split(" ")[0].replace(":", "")) + 1200
        )
        return courses

    
def main():
    timetable = "timetable.xlsx"
    ttparser = TimeTableParser(timetable)

    parsed_courses = ttparser.parse()

    json.dump(parsed_courses, open("timetable.json", "w"), indent=4)

if __name__ == "__main__":
    main()
